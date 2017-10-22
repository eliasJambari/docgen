import sys
import operator

from openpyxl import load_workbook

from excel import parser
from excel import tools
from excel import my_constants as cst


def map_columns_location(data_table):
    mapping = {}

    for entry in data_table:
        if len(entry) == 2:
            column = tools.letter_to_index(entry[cst.COLUMN])
            location = tools.check_location(entry[cst.LOCATION])

            # Check if not entry with same column/location
            if any(location in cells for cells in mapping.values()):
                sys.exit("Cannot have multiple entries with '" + location + "'")
            elif column in mapping.keys():
                mapping.get(column).append(location)
            else:
                mapping.update({column: [location]})
        else:
            sys.exit("No possible mapping with more than 2 columns")

    return mapping


def generate_excel_files(data_table, mapping, template_file, output_directory):
    template_file = tools.extract_one_sheet(template_file, output_directory, "template")

    columns = sorted(mapping.items(), key=operator.itemgetter(0))
    file_count = 0

    generated_files = []

    for element in data_table:
        file_count += 1

        new_file_name = tools.dupplicate_file(template_file, output_directory, str(file_count).rjust(3, '0') + ".xlsx")

        wb = load_workbook(filename=new_file_name)
        ws = wb["template"]

        fill_excel_file(ws, element, columns)

        wb.save(new_file_name)

        generated_files.append(new_file_name)

    return generated_files


def fill_excel_file(ws, content, columns):
    for i in range(len(content)):
        cell_content = content[i]
        location = columns[i]

        for cell in location[cst.LOCATION]:
            ws[cell] = cell_content


if __name__ == "__main__":
    file_name = sys.argv[1]
    sheet_name = sys.argv[2]

    sheets_info = {}

    sheets_info.update({sheet_name: [["A", "B"], 2]})

    data_tables = parser.parse_sheets(file_name, sheets_info)

    mapping = map_columns_location(data_tables[sheet_name])

    print(mapping)