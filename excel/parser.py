import sys
from openpyxl import load_workbook

from excel import tools
from excel import my_constants as cst


def parse_sheets(file_name, sheets_info, remove_empty_lines=True, plain_table=False):
    wb = load_workbook(filename=file_name, read_only=True)
    data_tables = {}

    for sheet_name, infos in sheets_info.items():
        ws = wb[sheet_name]

        # Prepare info of each worksheet
        columns = infos[cst.COLUMNS]
        first_row = infos[cst.FIRST_ROW]

        data_table = parse_sheet(ws, columns, first_row, remove_empty_lines)

        data_tables.update({sheet_name: data_table})

    if plain_table :
        if len(data_tables) == 1:
            return list(data_tables.values())[0]
        else:
            sys.exit("Plain table structure is only available when parsing one sheet")
    else:
        return data_tables


# Parse single sheet
def parse_sheet(ws, columns, first_row=2, remove_empty_lines=True):
    data = []

    col_indexes = tools.columns_to_indexes(columns, ws.max_column)

    # Iterate over each row of the sheet
    for row in ws.iter_rows(row_offset=first_row-1):
        row_data = []

        current_cell = 1
        next_cell = 0

        # Iterate over each cell of the current row
        for cell in row:
            if next_cell >= len(col_indexes):
                break
            if current_cell == col_indexes[next_cell]:
                row_data.append(tools.to_str(cell.value))
                next_cell += 1
            current_cell += 1

        # Add line if applicable
        if not(remove_empty_lines and tools.empty(row_data)):
            data.append(row_data)

    return data


if __name__ == "__main__":
    file_name = sys.argv[1]
    sheet_name = sys.argv[2]

    sheets_info = {}

    sheets_info.update({sheet_name: [["A", "B", "C", "D", "E"], 1]})

    data_tables = parse_sheets(file_name, sheets_info)

    tools.print_table(data_tables[sheet_name])